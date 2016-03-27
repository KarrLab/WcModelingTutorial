#!/usr/bin/python

'''
Generates the random components of the example model:
- Random RNA and protein species for each gene in the model
- Transcription, translation, and RNA degradation reactions for each RNA and protein species 
- Biomass production reaction (FBA objective) which is consistent with the specified biomass composition and transcription, translation, and RNA degradation reactions

@author Jonathan Karr, karr@mssm.edu
@date 3/24/2016
'''

#required libraries
from Bio import SeqIO
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from numpy import random
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
from util import N_AVOGADRO
import math
import numpy as np

#model parameters
cellVol = 4.58e-17 #volume (L)
concAa = 0.5e-3 #[ALA], [ARG], [CYS], ... (M)
concNxp = 1e-3 #[NTP], [NDP], [NMP] (M)
concPPi = 1e-3 #[PPi] (M)
concPi = 5e-3 #[Pi](M)
concH2O = 55 #[H2O] (M)
concH = 1.1220e-08 # [H] (M)
nRnaCopy = 5. #Copy numbers of RNA
nProtCopy = 100. #copy numbers of proteins
cellCycleLength = 8 * 60 * 60. #(s)
rnaHalfLife = 5 * 60. #(s)

#list of to generate random sequences for
genes = [
        {'id': 'RnaPolymerase', 'name': 'RNA polymerase'},
        {'id': 'Ribosome', 'name': 'Ribosome'},
        {'id': 'Rnase', 'name': 'RNase'},
        {'id': 'Adk', 'name': 'Adenylate kinase'},
        {'id': 'Apt', 'name': 'Adenine phosphoribosyltransferase '},
        {'id': 'Cmk', 'name': 'Cytidylate kinase'},
        {'id': 'Eno', 'name': 'Enolase'},
        {'id': 'Fba', 'name': 'Fructose-biphosphate aldolase'},
        {'id': 'Gap', 'name': 'Glyceraldehyde-3-phosphate-dehydrogenase'},
        {'id': 'Pgm', 'name': '2,3-bisphosphoglycerate-independent phosphoglycerate mutase '},
        {'id': 'Gk', 'name': 'Guanylate kinase'},
        {'id': 'Hpt', 'name': 'Hypoxanthine-guanine phosphoribosyltransferase'},
        {'id': 'Ldh', 'name': 'L-lactate dehydrogenase'},
        {'id': 'Nox', 'name': 'Probable NADH oxidase'},
        {'id': 'Pfk', 'name': '6-phosphofructokinase /Phosphohexokinase'},
        {'id': 'PgiB', 'name': 'Glucose-6-phosphate isomerase '},
        {'id': 'Pgk', 'name': 'Phosphoglycerate kinase'},
        {'id': 'Ppa', 'name': 'Inorganic pyrophosphatase '},
        {'id': 'Prs', 'name': 'Ribose-phosphate pyrophosphokinase '},
        {'id': 'Pyk', 'name': 'Pyruvate kinase'},
        {'id': 'Rpe', 'name': 'Probable ribulose-phosphate 3-epimerase'},
        {'id': 'LacA', 'name': 'Probable ribose-5-phosphate isomerase B '},
        {'id': 'Tim', 'name': 'Triosephosphate isomerase '},
        {'id': 'TklB', 'name': 'Transketolase '},
        {'id': 'Udk', 'name': 'Uridine kinase'},
        {'id': 'PyrH', 'name': 'Uridylate kinase '},
        {'id': 'Upp', 'name': 'Uracil phosphoribosyltransferase '},
        {'id': 'Pts', 'name': 'PTS EIIABC (glucose)'},
        {'id': 'PiAbcTransporter', 'name': 'Phosphate ABC transporter'},
        {'id': 'PeptAbcTransporter', 'name': 'Peptide ABC transporter'},
        ]

#1 and 3 letter amino acid codes
mwH = 1.0079
mwH2O = 18.0152
chargeH = 1
chargeH2O = 0

nmps = [
    {'id': 'A', 'name': 'AMP', 'molecularWeight': 345.205, 'charge': -2},
    {'id': 'C', 'name': 'CMP', 'molecularWeight': 321.18, 'charge': -2},
    {'id': 'G', 'name': 'GMP', 'molecularWeight': 361.204, 'charge': -2},
    {'id': 'U', 'name': 'UMP', 'molecularWeight': 322.165, 'charge': -2},
    ]
aminoAcids = [
    {'id': 'A', 'name': 'ALA', 'molecularWeight': 89.0929, 'charge': 0},
    {'id': 'R', 'name': 'ARG', 'molecularWeight': 175.208, 'charge': 1},
    {'id': 'D', 'name': 'ASP', 'molecularWeight': 132.094, 'charge': -1},
    {'id': 'N', 'name': 'ASN', 'molecularWeight': 132.118, 'charge': 0},
    {'id': 'C', 'name': 'CYS', 'molecularWeight': 121.158, 'charge': 0},
    {'id': 'Q', 'name': 'GLN', 'molecularWeight': 146.144, 'charge': 0},
    {'id': 'E', 'name': 'GLU', 'molecularWeight': 146.121, 'charge': -1},
    {'id': 'G', 'name': 'GLY', 'molecularWeight': 75.0664, 'charge': 0},
    {'id': 'H', 'name': 'HIS', 'molecularWeight': 155.154, 'charge': 0},
    {'id': 'I', 'name': 'ILE', 'molecularWeight': 131.172, 'charge': 0},
    {'id': 'L', 'name': 'LEU', 'molecularWeight': 131.172, 'charge': 0},
    {'id': 'K', 'name': 'LYS', 'molecularWeight': 147.195, 'charge': 1},
    {'id': 'M', 'name': 'MET', 'molecularWeight': 149.211, 'charge': 0},
    {'id': 'F', 'name': 'PHE', 'molecularWeight': 165.189, 'charge': 0},
    {'id': 'P', 'name': 'PRO', 'molecularWeight': 115.13, 'charge': 0},
    {'id': 'S', 'name': 'SER', 'molecularWeight': 105.092, 'charge': 0},
    {'id': 'T', 'name': 'THR', 'molecularWeight': 119.119, 'charge': 0},
    {'id': 'W', 'name': 'TRP', 'molecularWeight': 204.225, 'charge': 0},
    {'id': 'Y', 'name': 'TYR', 'molecularWeight': 181.188, 'charge': 0},
    {'id': 'V', 'name': 'VAL', 'molecularWeight': 117.146, 'charge': 0},
    ]

#1. Generates random RNA and protein species of length 3 * (protLen + 1) and protLen for each gene in the "genes" array. 
#2. Generates transcription, translation, and RNA degradation reactions for each RNA and protein species
#3. Calculates a biomass production reaction (FBA objective) which is consistent with the specified biomass composition (metabolite concentrations) and transcription, translation, and RNA degradation reactions.
#4. Outputs these species and reactions to an Excel file
def run(protLen = 100, startCodon = 'ATG', stopCodon = 'TAG'):
    random.seed(0)

    nGene = len(genes)
    
    '''generate random sequences'''
    dnaSeqs = []
    rnaSeqs = []
    protSeqs = []
    for gene in genes:
        dnaSeq = generateRandomSequence(protLen, startCodon, stopCodon)
        
        #create gene, RNA, and protein sequence records
        dnaSeqs.append(SeqRecord(dnaSeq, id = '%s-Gene' % gene['id'], description = '%s (Gene)' % gene['name']))
        rnaSeqs.append(SeqRecord(dnaSeq.transcribe(), id = '%s-Rna' % gene['id'], description = '%s (RNA)' % gene['name']))       
        protSeqs.append(SeqRecord(Seq(str(dnaSeq.translate(table=4))[0:-1]), id = '%s-Protein' % gene['id'], description = '%s (Protein)' % gene['name']))
        
    '''save sequences and reactions to Excel workbook'''
    wb = Workbook()
    
    specWs = wb.active
    specWs.title = "Species"
    rxnWs = wb.create_sheet(title='Reactions')
        
    #first rows
    specWs['A1'] = 'ID'
    specWs['B1'] = 'Name'
    specWs['C1'] = 'Structure'
    specWs['D1'] = 'Empirical formula'
    specWs['E1'] = 'Molecular weight'
    specWs['F1'] = 'Charge'
    specWs['G1'] = 'Type'
    specWs['H1'] = 'Subtype'
    specWs['I1'] = 'Average concentration, cytosol (M)'
    specWs['J1'] = 'Average concentration, extracellular space (M)'
    specWs['K1'] = 'Average count, cytosol'
    specWs['L1'] = 'Average count, extracellular space'
    specWs['M1'] = 'Cross reference source'
    specWs['N1'] = 'Cross reference ID'
    specWs['O1'] = 'Comments'
    
    rxnWs['A1'] = 'ID'
    rxnWs['B1'] = 'Name'
    rxnWs['C1'] = 'Submodel'
    rxnWs['D1'] = 'Stoichiometry'
    rxnWs['E1'] = 'Enzyme'
    rxnWs['F1'] = 'Rate law'
    rxnWs['G1'] = 'Vmax (1/s, 1/(M*s))'
    rxnWs['H1'] = 'Km (M)'
    rxnWs['I1'] = 'Cross reference source'
    rxnWs['J1'] = 'Cross reference ID'
    rxnWs['K1'] = 'Comments'
    
    for iGene, gene in enumerate(genes):
        #RNA species
        mw = 0
        charge = 0
        for nmp in nmps:
            mw += rnaSeqs[iGene].seq.count(nmp['id']) * nmp['molecularWeight']
            charge += rnaSeqs[iGene].seq.count(nmp['id']) * nmp['charge']
        l = len(rnaSeqs[iGene])
        mw = mw + (l - 1) * (mwH - mwH2O)
        charge = charge + (l - 1) * (chargeH - chargeH2O)
        
        specWs['A%d' % (iGene + 2)] = '%s-Rna' % gene['id'] #ID
        specWs['B%d' % (iGene + 2)] = '%s (RNA)' % gene['name'] #name
        specWs['C%d' % (iGene + 2)] = str(rnaSeqs[iGene].seq) #Structure (sequence)
        specWs['D%d' % (iGene + 2)] = '' #Empirical formula
        specWs['E%d' % (iGene + 2)] = mw #molecular weight
        specWs['F%d' % (iGene + 2)] = charge #charge
        specWs['G%d' % (iGene + 2)] = 'RNA' #type
        specWs['H%d' % (iGene + 2)] = '' #subtype
        specWs['I%d' % (iGene + 2)] =  nRnaCopy / cellVol / N_AVOGADRO #cytosol concentration
        specWs['J%d' % (iGene + 2)] =  0 #extracellular concentration
        specWs['K%d' % (iGene + 2)] =  nRnaCopy #cytosol count
        specWs['L%d' % (iGene + 2)] =  0 #extracellular count
        specWs['M%d' % (iGene + 2)] = '' #Cross reference source
        specWs['N%d' % (iGene + 2)] = '' #Cross reference ID
        specWs['O%d' % (iGene + 2)] = '' #Comments
        
        #protein species
        mw = 0
        charge = 0
        for aa in aminoAcids:
            mw += protSeqs[iGene].seq.count(aa['id']) * aa['molecularWeight']
            charge += protSeqs[iGene].seq.count(aa['id']) * aa['charge']
        l = len(protSeqs[iGene])
        mw -= (l - 1) * mwH2O
        charge -= (l - 1) * chargeH2O
        
        specWs['A%d' % (iGene + 2 + nGene)] = '%s-Protein' % gene['id'] #ID
        specWs['B%d' % (iGene + 2 + nGene)] = '%s (Protein)' % gene['name'] #name
        specWs['C%d' % (iGene + 2 + nGene)] = str(protSeqs[iGene].seq) #Structure
        specWs['D%d' % (iGene + 2 + nGene)] = '' #Empirical formula
        specWs['E%d' % (iGene + 2 + nGene)] = mw #molecular weight
        specWs['F%d' % (iGene + 2 + nGene)] = charge #charge
        specWs['G%d' % (iGene + 2 + nGene)] = 'Protein' #type
        specWs['H%d' % (iGene + 2 + nGene)] = '' #subtype
        specWs['I%d' % (iGene + 2 + nGene)] =  nProtCopy / cellVol / N_AVOGADRO #cytosol concentration
        specWs['J%d' % (iGene + 2 + nGene)] =  0 #extracellular concentration
        specWs['K%d' % (iGene + 2 + nGene)] =  nProtCopy #cytosol count
        specWs['L%d' % (iGene + 2 + nGene)] =  0 #extracellular count
        specWs['M%d' % (iGene + 2 + nGene)] = '' #Cross reference source
        specWs['N%d' % (iGene + 2 + nGene)] = '' #Cross reference ID
        specWs['O%d' % (iGene + 2 + nGene)] = '' #Comments
        
        #transcription reaction
        nA = rnaSeqs[iGene].seq.count('A')
        
        rxnWs['A%d' % (iGene + 2)] = 'Transcription-%s' % gene['id'] #ID
        rxnWs['B%d' % (iGene + 2)] = 'Transcription (%s)' % gene['name'] #name
        rxnWs['C%d' % (iGene + 2)] = 'Transcription' #submodel
        rxnWs['D%d' % (iGene + 2)] = '[c]: (%d) ATP + (%d) CTP + (%d) GTP + (%d) UTP + H2O ==> %s-Rna + (%d) PPI + H' % ( #stoichometry
            rnaSeqs[iGene].seq.count('A'),
            rnaSeqs[iGene].seq.count('C'),
            rnaSeqs[iGene].seq.count('G'),
            rnaSeqs[iGene].seq.count('U'),
            gene['id'],
            len(rnaSeqs[iGene].seq), )
        rxnWs['E%d' % (iGene + 2)] = 'RnaPolymerase-Protein' #enzyme
        rxnWs['F%d' % (iGene + 2)] = 'Vmax * min(ATP[c], CTP[c], GTP[c], UTP[c]) / (Km + min(ATP[c], CTP[c], GTP[c], UTP[c])) * RnaPolymerase-Protein[c]' #rate law
        rxnWs['G%d' % (iGene + 2)] = math.log(2) * (1 / rnaHalfLife + 1 / cellCycleLength) * 2 * nRnaCopy / nProtCopy #Vmax
        rxnWs['H%d' % (iGene + 2)] = concNxp #Km
        rxnWs['I%d' % (iGene + 2)] = 'EC' #Cross reference source
        rxnWs['J%d' % (iGene + 2)] = '2.7.7.6' #Cross reference ID
        rxnWs['K%d' % (iGene + 2)] = 'Kinetics constrained by cell theory (mass doubling).' #Comments
        
        #translation reaction
        aaCnts = []
        for aa in aminoAcids:
            aaCnts.append('(%d) %s' % (protSeqs[iGene].seq.count(aa['id']), aa['name'], ))
            
        rxnWs['A%d' % (iGene + 2 + nGene)] = 'Translation-%s' % gene['id'] #ID
        rxnWs['B%d' % (iGene + 2 + nGene)] = 'Translation (%s)' % gene['name'] #name
        rxnWs['C%d' % (iGene + 2 + nGene)] = 'Translation' #submodel
        rxnWs['D%d' % (iGene + 2 + nGene)] = '[c]: %s + (%d) GTP + (%d) H2O ==> %s-Protein + (%d) GDP + (%d) PI + (%d) H' % ( #stoichometry
            ' + '.join(aaCnts),
            2 * len(protSeqs[iGene].seq) + 3,
            2 * len(protSeqs[iGene].seq) + 3 - (len(protSeqs[iGene].seq) - 1),
            gene['id'],
            2 * len(protSeqs[iGene].seq),
            2 * len(protSeqs[iGene].seq),
            2 * len(protSeqs[iGene].seq),)
        rxnWs['E%d' % (iGene + 2 + nGene)] = 'Ribosome-Protein' #enzyme
        rxnWs['F%d' % (iGene + 2 + nGene)] = 'Vmax * min(ALA[c], ARG[c], ASP[c], ASN[c], CYS[c], GLN[c], GLU[c], GLY[c], HIS[c], ILE[c], LEU[c], LYS[c], MET[c], PHE[c], PRO[c], SER[c], THR[c], TRP[c], TYR[c], VAL[c]) / (Km + min(ALA[c], ARG[c], ASP[c], ASN[c], CYS[c], GLN[c], GLU[c], GLY[c], HIS[c], ILE[c], LEU[c], LYS[c], MET[c], PHE[c], PRO[c], SER[c], THR[c], TRP[c], TYR[c], VAL[c])) * %s-Rna[c] * Ribosome-Protein[c]' % gene['id'] #rate law
        rxnWs['G%d' % (iGene + 2 + nGene)] = math.log(2) / cellCycleLength * 2 / nRnaCopy #Vmax
        rxnWs['H%d' % (iGene + 2 + nGene)] = concAa #Km
        rxnWs['I%d' % (iGene + 2 + nGene)] = 'EC' #Cross reference source
        rxnWs['J%d' % (iGene + 2 + nGene)] = '2.3.2.12' #Cross reference ID
        rxnWs['K%d' % (iGene + 2 + nGene)] = 'Kinetics constrained by cell theory (mass doubling).' #Comments
        
        #RNA degradation reaction
        rxnWs['A%d' % (iGene + 2 + nGene * 2)] = 'RnaDegradation-%s' % gene['id'] #ID
        rxnWs['B%d' % (iGene + 2 + nGene * 2)] = 'RNA degradation (%s)' % gene['name'] #name
        rxnWs['C%d' % (iGene + 2 + nGene * 2)] = 'RnaDegradation' #submodel
        rxnWs['D%d' % (iGene + 2 + nGene * 2)] = '[c]: %s-Rna + (%d) H2O ==> (%d) AMP + (%d) CMP + (%d) GMP + (%d) UMP + (%d) H' % ( #stoichometry
            gene['id'],
            len(rnaSeqs[iGene].seq) - 1,
            rnaSeqs[iGene].seq.count('A'), 
            rnaSeqs[iGene].seq.count('C'), 
            rnaSeqs[iGene].seq.count('G'),
            rnaSeqs[iGene].seq.count('U'),
            len(rnaSeqs[iGene].seq) - 1, )
        rxnWs['E%d' % (iGene + 2 + nGene * 2)] = 'Rnase-Protein' #enzyme
        rxnWs['F%d' % (iGene + 2 + nGene * 2)] = 'Vmax * %s-Rna[c] / (%s-Rna[c] + Km) * Rnase-Protein[c]' % (gene['id'], gene['id']) #rate law
        rxnWs['G%d' % (iGene + 2 + nGene * 2)] = math.log(2) / rnaHalfLife * 2 * nRnaCopy / nProtCopy #Vmax
        rxnWs['H%d' % (iGene + 2 + nGene * 2)] = nRnaCopy / cellVol / N_AVOGADRO #Km
        rxnWs['I%d' % (iGene + 2 + nGene * 2)] = 'EC' #Cross reference source
        rxnWs['J%d' % (iGene + 2 + nGene * 2)] = '3.1.-.-' #Cross reference ID
        rxnWs['K%d' % (iGene + 2 + nGene * 2)] = 'Kinetics determined by RNA half liife.' #Comments
        
    #biomass production
    baseCntsNtp = {'A': 0, 'C': 0, 'G': 0, 'U': 0}
    nNtp_trn = {'A': 0, 'C': 0, 'G': 0, 'U': 0}
    nNmp_dcy = {'A': 0, 'C': 0, 'G': 0, 'U': 0}
    nAa_trl = {}
    for aa in aminoAcids:
        nAa_trl[aa['id']] = 0
    for iGene, gene in enumerate(genes):
        baseCntsNtp['A'] += rnaSeqs[iGene].seq.count('A') * nRnaCopy
        baseCntsNtp['C'] += rnaSeqs[iGene].seq.count('C') * nRnaCopy
        baseCntsNtp['G'] += rnaSeqs[iGene].seq.count('G') * nRnaCopy
        baseCntsNtp['U'] += rnaSeqs[iGene].seq.count('U') * nRnaCopy
        for aa in aminoAcids:
            nAa_trl[aa['id']] += protSeqs[iGene].seq.count(aa['id']) * nProtCopy
                
    nNtp_trn['A'] = baseCntsNtp['A'] * (1 + cellCycleLength / rnaHalfLife)
    nNtp_trn['C'] = baseCntsNtp['C'] * (1 + cellCycleLength / rnaHalfLife)
    nNtp_trn['G'] = baseCntsNtp['G'] * (1 + cellCycleLength / rnaHalfLife)
    nNtp_trn['U'] = baseCntsNtp['U'] * (1 + cellCycleLength / rnaHalfLife)
    
    nNmp_dcy['A'] = baseCntsNtp['A'] * cellCycleLength / rnaHalfLife
    nNmp_dcy['C'] = baseCntsNtp['C'] * cellCycleLength / rnaHalfLife
    nNmp_dcy['G'] = baseCntsNtp['G'] * cellCycleLength / rnaHalfLife
    nNmp_dcy['U'] = baseCntsNtp['U'] * cellCycleLength / rnaHalfLife
    
    nPPi_trn = (protLen + 1) * 3 * nRnaCopy * len(genes) * (1 + cellCycleLength / rnaHalfLife)
    nH2O_trn = nRnaCopy * len(genes) * (1 + cellCycleLength / rnaHalfLife)
    nH2O_dcy = ((protLen + 1) * 3 - 1) * nRnaCopy * len(genes) * (cellCycleLength / rnaHalfLife)
   
    nGtp_trl = (2 * protLen + 3) * nProtCopy * len(genes)
    nH2o_trl = (protLen + 4) * nProtCopy * len(genes)
        
    stochArr = []
    stochArr.append({'species': 'Biomass', 'coeff': 1})
    
    stochArr.append({'species': 'H2O', 'coeff': 0
        - cellVol * concH2O * N_AVOGADRO 
        - nH2O_trn 
        - nH2o_trl
        - nH2O_dcy
        })
    stochArr.append({'species': 'H', 'coeff': 0
        - cellVol * concH * N_AVOGADRO
        + nH2O_trn
        + nGtp_trl
        + nH2O_dcy
        })
        
    stochArr.append({'species': 'ATP', 'coeff': -cellVol * concNxp * N_AVOGADRO - nNtp_trn['A']})
    stochArr.append({'species': 'CTP', 'coeff': -cellVol * concNxp * N_AVOGADRO - nNtp_trn['C']})
    stochArr.append({'species': 'GTP', 'coeff': -cellVol * concNxp * N_AVOGADRO - nNtp_trn['G'] - nGtp_trl})
    stochArr.append({'species': 'UTP', 'coeff': -cellVol * concNxp * N_AVOGADRO - nNtp_trn['U']})
    
    stochArr.append({'species': 'GDP', 'coeff': -cellVol * concNxp * N_AVOGADRO + nGtp_trl})
    
    stochArr.append({'species': 'AMP', 'coeff': -cellVol * concNxp * N_AVOGADRO + nNmp_dcy['A']})
    stochArr.append({'species': 'CMP', 'coeff': -cellVol * concNxp * N_AVOGADRO + nNmp_dcy['C']})
    stochArr.append({'species': 'GMP', 'coeff': -cellVol * concNxp * N_AVOGADRO + nNmp_dcy['G']})
    stochArr.append({'species': 'UMP', 'coeff': -cellVol * concNxp * N_AVOGADRO + nNmp_dcy['U']})
    
    stochArr.append({'species': 'PPI', 'coeff': -cellVol * concPPi * N_AVOGADRO + nPPi_trn})
    stochArr.append({'species': 'PI', 'coeff': -cellVol * concPi * N_AVOGADRO + nGtp_trl})
    
    for aa in aminoAcids:
        stochArr.append({'species': aa['name'], 'coeff': -cellVol * concAa * N_AVOGADRO - nAa_trl[aa['id']]})
    
    lhs = []
    rhs = []
    for part in stochArr:
        if part['coeff'] < 0:
            lhs.append('(%e) %s' % (-part['coeff'], part['species']))
        else:
            rhs.append('(%e) %s' % ( part['coeff'], part['species']))
    
    stoichiometry = '[c]: %s ==> %s' % (' + '.join(lhs), ' + '.join(rhs))
        
    rxnWs['A%d' % (nGene * 3 + 2)] = 'BiomassProduction' #id
    rxnWs['B%d' % (nGene * 3 + 2)] = 'Biomass production' #id
    rxnWs['C%d' % (nGene * 3 + 2)] = 'Metabolism' #id
    rxnWs['D%d' % (nGene * 3 + 2)] = stoichiometry #stoichiometry
    rxnWs['E%d' % (nGene * 3 + 2)] = '' #enzyme
    rxnWs['F%d' % (nGene * 3 + 2)] = '' #rate law
    rxnWs['G%d' % (nGene * 3 + 2)] = '' #Vmax
    rxnWs['H%d' % (nGene * 3 + 2)] = '' #Km
    rxnWs['I%d' % (nGene * 3 + 2)] = '' #Cross reference source
    rxnWs['J%d' % (nGene * 3 + 2)] = '' #Cross reference id
    
    #style
    specWs.freeze_panes = specWs['B2']
    rxnWs.freeze_panes = rxnWs['B2']   
    
    for i in range(1, 15 + 1):
        cell = specWs.cell(row=1, column=i)
        cell.font = cell.font.copy(bold = True)
        cell.fill = cell.fill.copy(fgColor=Color('CCCCCC'), patternType='solid')
        
    for i in range(1, 11 + 1):
        cell = rxnWs.cell(row=1, column=i)
        cell.font = cell.font.copy(bold = True)
        cell.fill = cell.fill.copy(fgColor=Color('CCCCCC'), patternType='solid')
        
    for i in range(1, 2 * nGene + 2):
        specWs.row_dimensions[i].height = 15
        for j in range(1, 15 + 1):
            cell = specWs.cell(row=i, column=j)
            cell.alignment = cell.alignment.copy(wrap_text = True)
            
    for i in range(1, 3 * nGene + 3):
        rxnWs.row_dimensions[i].height = 15
        for j in range(1, 11 + 1):
            cell = rxnWs.cell(row=i, column=j)
            cell.alignment = cell.alignment.copy(wrap_text = True)
    
    wb.save("data/Model-RNA and proteins.xlsx")    

    '''save RNA and protein sequences in FASTA format'''
    output_handle = open("data/Genes.fasta", "w")
    SeqIO.write(rnaSeqs, output_handle, "fasta")
    output_handle.close()
    
    output_handle = open("data/RNAs.fasta", "w")
    SeqIO.write(rnaSeqs, output_handle, "fasta")
    output_handle.close()
    
    output_handle = open("data/Proteins.fasta", "w")
    SeqIO.write(protSeqs, output_handle, "fasta")
    output_handle.close()
    
#Generates random protein-coding DNA sequence of length 3 * (protLen + 1) that begins with a start codon and ends with a stop codon such that the corresponding protein has length protLen.
def generateRandomSequence(protLen = 100, startCodon = 'ATG', stopCodon = 'TAG'):
    dnaBases = 'ACGT'
    
    dnaSeqStr = startCodon
    for iCodon in range(protLen - 1):
        validCodon = False
        while not(validCodon):
            codonStr = '' \
                + dnaBases[random.randint(0, 4)] \
                + dnaBases[random.randint(0, 4)] \
                + dnaBases[random.randint(0, 4)]
            validCodon = Seq(dnaSeqStr + codonStr).translate(table=4).find('*') == -1
        dnaSeqStr += codonStr
    dnaSeqStr += stopCodon
        
    return Seq(dnaSeqStr)
    
if __name__ == "__main__":
    run()
