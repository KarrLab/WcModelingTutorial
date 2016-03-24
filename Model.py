''' 
Reads models specified in Excel into a Python object

Author: Jonathan Karr, karr@mssm.edu
Last updated: 3/22/2016
'''

#required libraries
from cobra import Metabolite as CobraMetabolite
from cobra import Model as CobraModel
from cobra import Reaction as CobraReaction
from openpyxl import load_workbook
from pprint import PrettyPrinter
import math
import numpy as np
import re

#Represents a model (submodels, compartments, species, reactions, parameters, references)
class Model:
    submodels = []
    compartments = []
    species = []
    reactions = []
    parameters = []
    references = []
    
    speciesCounts = np.zeros((0, 0)) #rows: species, columns: compartments
    
    def __init__(self, submodels = [], compartments = [], species = [], reactions = [], parameters = [], references = []):
        self.submodels = submodels
        self.compartments = compartments
        self.species = species
        self.reactions = reactions
        self.parameters = parameters
        self.references = references
        
    def setupForSimulation(self):
        self.speciesCounts = np.zeros((len(self.species), len(self.compartments)))
        
        for submdl in self.submodels:
            submdl.setupForSimulation()
        
    def getComponentById(self, id):
        components = self.submodels + self.compartments + self.species + self.reactions + self.parameters + self.references
        for component in components:
            if component.id == id:
                return component
               
#Represents a submodel
class Submodel:
    id = ''
    name = ''
    algorithm = ''
    reactions = []
    reactionParticipants = []
    enzymes = []
    
    reactionParticipantCounts = np.zeros((0, 0))
    enzymeCounts = np.zeros((0, 0))
    
    def __init__(self, id = '', name = '', reactions = [], reactionParticipants = [], enzymeCompartments = []):
        self.id = id
        self.name = name
        self.reactions = reactions
        self.reactionParticipants = reactionParticipants
        self.enzymeCompartments = enzymeCompartments
    
    def setupForSimulation(self):
        self.reactionParticipantCounts = np.zeros((len(self.reactionParticipants), 1))
        self.enzymeCounts = np.zeros((len(self.enzymes), 1))
        
#Represents an FBA submodel
class FbaSubmodel(Submodel):
    cobraModel = None
    
    def __init__(self, *args, **kwargs):        
        Submodel.__init__(self, *args, **kwargs)
        self.algorithm = 'FBA'
        
    def setupForSimulation(self):
        '''setup reaction participant, enzyme counts matrices'''
        Submodel.setupForSimulation(self)
                
        '''Setup FBA'''
        self.cobraModel = CobraModel(self.id)
            
        #setup metabolites
        cbMets = []
        for part in self.reactionParticipants:
            cbMets.append(CobraMetabolite(
                id = '%s[%s]' % (part.species.id, part.compartment.id),
                name = '%s (%s)' % (part.species.name, part.compartment.name)))
        self.cobraModel.add_metabolites(cbMets)
        
        #setup reactions
        for rxn in self.reactions:            
            cbRxn = CobraReaction(
                id = rxn.id,
                name = rxn.name,
                lower_bound = -1000. if rxn.reversible else 0,
                upper_bound = 1000.,
                objective_coefficient = 1. if rxn.id == 'BiomassProduction' else 0.,
                )
            self.cobraModel.add_reaction(cbRxn)

            cbMets = {}
            for part in rxn.participants:
                cbMets['%s[%s]' % (part.species.id, part.compartment.id)] = part.coefficient
            cbRxn.add_metabolites(cbMets)            
        
        #add external exchange reactions
        for part in self.reactionParticipants:
            if part.compartment.id == 'e':
                cbRxn = CobraReaction(
                    id = '%sEx' % part.species.id,
                    name = '%s exchange' % part.species.name,
                    lower_bound = -1000.,
                    upper_bound = 1000.,
                    objective_coefficient = 0.,
                    )
                self.cobraModel.add_reaction(cbRxn)
                cbRxn.add_metabolites({'%s[e]' % part.species.id: 1})
        
        #add biomass exchange reaction
        cbRxn = CobraReaction(
            id = 'BiomassEx',
            name = 'Biomass exchange',
            lower_bound = 0.,
            upper_bound = 1000.,
            objective_coefficient = 0.,
            )
        self.cobraModel.add_reaction(cbRxn)
        cbRxn.add_metabolites({'Biomass[c]': -1})
        
        cbRxn = CobraReaction(
            id = 'BiomassEx2',
            name = 'Biomass exchange 2',
            lower_bound = 0.,
            upper_bound = 1000.,
            objective_coefficient = 0.,
            )
        self.cobraModel.add_reaction(cbRxn)
        cbRxn.add_metabolites({'Biomass[c]': 1})
        
#Represents an SSA submodel
class SsaSubmodel(Submodel):
    def __init__(self, *args, **kwargs):
        Submodel.__init__(self, *args, **kwargs)
        self.algorithm = 'SSA'
        
    def setupForSimulation(self):
        Submodel.setupForSimulation(self)
        
#Represents a compartment
class Compartment:
    id = ''
    name = ''
    volume = None
    comments = ''
    
    def __init__(self, id = '', name = '', volume = None, comments = ''):
        self.id = id
        self.name = name
        self.volume = volume
        self.comments = comments
        
#Represents a species
class Species:
    id = ''
    name = ''
    structure = ''
    charge = None
    type = ''
    subtype = ''
    concentrations = []
    crossRefs = []
    
    def __init__(self, id = '', name = '', structure = '', charge = None, type = '', subtype = '', concentrations = [], crossRefs = []): 
        self.id = id    
        self.name = name
        self.structure = structure
        self.charge = charge
        self.type = type
        self.subtype = subtype
        self.concentrations = concentrations
        self.crossRefs = crossRefs

#Represents a reaction
class Reaction:
    id = ''
    name = ''
    submodel = ''
    reversible = None
    participants = []
    enzyme = ''
    rateLaw = None
    vmax = None
    km = None
    crossRefs = []

    def __init__(self, id = '', name = '', submodel = '', reversible = None, participants = [], enzyme = '', rateLaw = '', vmax = None, km = None, crossRefs = []):        
        self.id = id    
        self.name = name
        self.submodel = submodel
        self.reversible = reversible
        self.participants = participants
        self.enzyme = enzyme
        self.rateLaw = rateLaw
        self.vmax = vmax
        self.km = km
        self.crossRefs = crossRefs
        
    def getStoichiometryString(self):
        globalComp = self.participants[0].compartment
        for part in self.participants:
            if part.compartment != globalComp:
                globalComp = None
                break
        
        lhs = []
        rhs = []
        for part in self.participants:
            if part.coefficient < 0:
                partStr = ''
                if part.coefficient != -1:
                    if math.ceil(part.coefficient) == part.coefficient: 
                        partStr += '(%d) ' % -part.coefficient
                    else:
                        partStr += '(%e) ' % -part.coefficient
                partStr += part.species.id
                if globalComp is None:
                    partStr += '[%s]' % part.compartment.id
                lhs.append(partStr)
            else:
                partStr = ''
                if part.coefficient != 1:
                    if math.ceil(part.coefficient) == part.coefficient: 
                        partStr += '(%d) ' % part.coefficient
                    else:
                        partStr += '(%e) ' % part.coefficient
                partStr += part.species.id
                if globalComp is None:
                    partStr += '[%s]' % part.compartment.id
                rhs.append(partStr)
            
        stoichStr = ''
        if globalComp is not None:
            stoichStr += '[%s]: ' % globalComp.id
        stoichStr += '%s %s==> %s' % (' + '.join(lhs), '<' if self.reversible else '', ' + '.join(rhs))
        
        return stoichStr
        
#Represents a model parameter
class Parameter:
    id = ''
    name = ''
    value = None
    units = ''
    comments = ''
    
    def __init__(self, id = '', name = '', value = None, units = '', comments = ''):
        self.id = id
        self.name = name
        self.value = value
        self.units = units
        self.comments = comments
            
#Represents a reference
class Reference:
    id = ''
    name = ''
    crossRefs = []
    comments = ''
    
    def __init__(self, id = '', name = '', crossRefs = [], comments = ''):
        self.id = id
        self.name = name
        self.crossRefs = crossRefs
        self.comments = comments

#Represents a concentration in a compartment
class Concentration:
    compartment = ''
    value = None
    
    def __init__(self, compartment = '', value = None):
        self.compartment = compartment
        self.value = value
    
#Represents a participant in a reaction
class SpeciesCompartment:
    species = ''
    compartment = ''
    
    def __init__(self, species = '', compartment = ''):
        self.species = species
        self.compartment = compartment    
    
#Represents a participant in a reaction
class ReactionParticipant:
    species = ''
    compartment = ''
    coefficient = None
    
    def __init__(self, species = '', compartment = '', coefficient = None):
        self.species = species
        self.compartment = compartment    
        self.coefficient = coefficient
    
#Represents a cross reference to an external database
class CrossReference:
    source = ''
    id = ''    

    def __init__(self, source = '', id = ''):
        self.source = source
        self.id = id

#Reads model from Excel file into a Python object
def getModelFromExcel(filename):
    wb = load_workbook(filename = filename)

    #initialize model object
    mdl = Model()

    '''Read details from Excel'''
    #submodels
    ws = wb['Submodels']
    for iRow in range(2, ws.max_row + 1):
        id = ws.cell(row = iRow, column = 1).value
        name = ws.cell(row = iRow, column = 2).value
        algorithm = ws.cell(row = iRow, column = 3).value
        if algorithm == 'FBA':
            submdl = FbaSubmodel(id = id, name = name)
        elif algorithm == 'SSA':
            submdl = SsaSubmodel(id = id, name = name)
        else:
            raise Exception('Undefined algorithm "%s" for submodel "%s"' % (algorithm, id))
        mdl.submodels.append(submdl)
            
    #compartments
    ws = wb['Compartments']
    for iRow in range(2, ws.max_row + 1):
        mdl.compartments.append(Compartment(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            volume = float(ws.cell(row = iRow, column = 3).value),
            comments = ws.cell(row = iRow, column = 4).value,
            ))
    
    #species
    ws = wb['Species']
    for iRow in range(2, ws.max_row + 1):
        chargeStr = ws.cell(row = iRow, column = 4).value
        if chargeStr:
            charge = float(chargeStr)
        else:
            charge = None
    
        mdl.species.append(Species(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            structure = ws.cell(row = iRow, column = 3).value,
            charge = charge,
            type = ws.cell(row = iRow, column = 5).value,
            subtype = ws.cell(row = iRow, column = 6).value, 
            concentrations = [
                Concentration(compartment = 'c', value = float(ws.cell(row = iRow, column = 7).value or 0)),
                Concentration(compartment = 'e', value = float(ws.cell(row = iRow, column = 8).value or 0)),
                ],
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 11).value, 
                    id = ws.cell(row = iRow, column = 12).value,
                    ),
                ],
            ))
            
    #reactions
    ws = wb['Reactions']
    for iRow in range(2, ws.max_row + 1):
        stoichiometry = parseStoichiometry(ws.cell(row = iRow, column = 4).value)
        mdl.reactions.append(Reaction(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            submodel = ws.cell(row = iRow, column = 3).value,
            reversible = stoichiometry['reversible'],
            participants = stoichiometry['participants'],
            enzyme = ws.cell(row = iRow, column = 5).value,
            rateLaw = ws.cell(row = iRow, column = 6).value,
            vmax = ws.cell(row = iRow, column = 7).value,
            km = ws.cell(row = iRow, column = 8).value,
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 9).value, 
                    id = ws.cell(row = iRow, column = 10).value,
                    ),
                ],
            ))
            
    #parameters
    ws = wb['Parameters']
    for iRow in range(2, ws.max_row + 1):
        mdl.parameters.append(Parameter(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            value = float(ws.cell(row = iRow, column = 3).value),
            units = ws.cell(row = iRow, column = 4).value,
			comments = ws.cell(row = iRow, column = 5).value,
            ))
            
    #references
    ws = wb['References']
    for iRow in range(2, ws.max_row + 1):
        mdl.references.append(Reference(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 3).value, 
                    id = ws.cell(row = iRow, column = 4).value,
                    ),
                ],
            comments = ws.cell(row = iRow, column = 5).value,
            ))
            
    '''deserialize references'''
    undefinedComponents = []
    
    for species in mdl.species:
        for conc in species.concentrations:
            id = conc.compartment            
            obj = mdl.getComponentById(id)
            if id and obj is None:
                undefinedComponents.append(id)
            conc.compartment = obj                
            
    for reaction in mdl.reactions:
        id = reaction.submodel        
        obj = mdl.getComponentById(id)
        if id and obj is None:
            undefinedComponents.append(id)
        reaction.submodel = obj
        
        for part in reaction.participants:
            id = part.species            
            obj = mdl.getComponentById(id)
            if id and obj is None:               
                undefinedComponents.append(id)
            part.species = obj
                
            id = part.compartment            
            obj = mdl.getComponentById(id)
            if id and obj is None:
                undefinedComponents.append(id)
            part.compartment = obj
        
        id = reaction.enzyme
        obj = mdl.getComponentById(id)
        if id and obj is None:
            undefinedComponents.append(id)
        reaction.enzyme = obj

    if len(undefinedComponents) > 0:
        undefinedComponents = list(set(undefinedComponents))
        undefinedComponents.sort()
        raise Exception('Undefined components:\n- %s' % ('\n- '.join(undefinedComponents)))
        
    for submdl in mdl.submodels:
        submdl.reactions = []
        submdl.reactionParticipants = []
        submdl.enzymes = []
    for rxn in mdl.reactions:
        rxn.submodel.reactions.append(rxn)
        for part in rxn.participants:
            rxn.submodel.reactionParticipants.append('%s[%s' % (part.species.id, part.compartment.id))
        if rxn.enzyme:
            rxn.submodel.enzymeCompartments.append(rxn.enzyme.id)
    for submdl in mdl.submodels:
        partStrArr = list(set(submdl.reactionParticipants))
        partStrArr.sort()
        submdl.reactionParticipants = []
        for partStr in partStrArr:
            speciesId, compId = partStr.split('[')
            submdl.reactionParticipants.append(SpeciesCompartment(
                species = mdl.getComponentById(speciesId),
                compartment = mdl.getComponentById(compId),
                ))
        
        enzStrArr = list(set(submdl.enzymes))
        enzStrArr.sort()
        submdl.enzymes = []
        for enzStr in enzStrArr:
            submdl.enzymes.append(SpeciesCompartment(
                species = mdl.getComponentById(enzStr), 
                compartment = mdl.getComponentById('c'),
                ))
        
    '''Prepare submodels for computation'''
    mdl.setupForSimulation()    
        
    '''Return'''
    return mdl
    
#Parse a string representing the stoichiometry of a reaction into a Python object
def parseStoichiometry(rxnStr):
    #Split stoichiometry in to global compartment, left-hand side, right-hand side, reversibility indictor
	rxnMatch = re.match('(?P<compartment>\[([a-z])\]: )?(?P<lhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?) (?P<direction>[<]?)==> (?P<rhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?)', rxnStr, flags=re.I)
	if rxnMatch is None:
		raise Exception('Invalid stoichiometry: %s' % rxnStr)
		
    #Determine reversiblity
	rxnDict = rxnMatch.groupdict()
	reversible = rxnDict['direction'] == '<'
	
    #Determine if global compartment for reaction was specified
	if rxnDict['compartment'] is None:
		globalComp = None
	else:
		globalComp = re.match('\[(?P<compartment>[a-z])\]', rxnDict['compartment'], flags=re.I).groupdict()['compartment']
	
    #initialize array of reaction participants
	participants = []
	
    #Parse left-hand side
	for rxnPartStr in rxnDict['lhs'].split(' + '):
		rxnPartDict = re.match('(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxnPartStr, flags=re.I).groupdict()
				
		species = rxnPartDict['species']
		compartment = rxnPartDict['compartment'] or globalComp
		coefficient = float(rxnPartDict['coefficient'] or 1)
	
		participants.append(ReactionParticipant(
			species = species,
			compartment = compartment,
			coefficient = -coefficient,
			))
			
    #Parse right-hand side
	for rxnPartStr in rxnDict['rhs'].split(' + '):
		rxnPartDict = re.match('(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxnPartStr, flags=re.I).groupdict()
				
		species = rxnPartDict['species']
		compartment = rxnPartDict['compartment'] or globalComp
		coefficient = float(rxnPartDict['coefficient'] or 1)
	
		participants.append(ReactionParticipant(
			species = species,
			compartment = compartment,
			coefficient = coefficient,
			))

	return {
		'reversible': reversible,
		'participants': participants,
		}