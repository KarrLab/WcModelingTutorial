''' 
Reads models specified in Excel into a Python object

@author Jonathan Karr, karr@mssm.edu
@date 3/22/2016
'''

#required libraries
from cobra import Metabolite as CobraMetabolite
from cobra import Model as CobraModel
from cobra import Reaction as CobraReaction
from itertools import chain
from numpy import random
from openpyxl import load_workbook
from util import N_AVOGADRO
import math
import numpy as np
import util
import re
import warnings

#Represents a model (submodels, compartments, species, reactions, parameters, references)
class Model:
    submodels = []
    compartments = []
    species = []
    reactions = []
    parameters = []
    references = []
    
    density = None
    fractionDryWeight = None
    
    speciesCounts = np.zeros(0) #rows: species, columns: compartments
    mass = None #cell mass
    dryWeight = None #cell dry weight
    volume = None #cell volume
    extracellularVolume = None #media volume
    growth = None
    
    def __init__(self, submodels = [], compartments = [], species = [], reactions = [], parameters = [], references = []):
        self.submodels = submodels
        self.compartments = compartments
        self.species = species
        self.reactions = reactions
        self.parameters = parameters
        self.references = references       
        
    def setupSimulation(self):
        self.fractionDryWeight = self.getComponentById('fractionDryWeight', self.parameters).value
    
        for subModel in self.submodels:
            subModel.setupSimulation()

        self.calcInitialConditions()
            
    def calcInitialConditions(self):
        cellComp = self.getComponentById('c', self.compartments)
        extrComp = self.getComponentById('e', self.compartments)
        
        #volume
        self.volume = cellComp.initialVolume
        self.extracellularVolume = extrComp.initialVolume
        
        #species counts
        self.speciesCounts = np.zeros((len(self.species), len(self.compartments)))
        for species in self.species:
            for conc in species.concentrations:
                self.speciesCounts[species.index, conc.compartment.index] = conc.value * conc.compartment.initialVolume * N_AVOGADRO
        
        #cell mass
        self.calcMass()                
         
        #density
        self.density = self.mass / self.volume
        
        #growth
        self.growth = np.nan
        
        #sync submodels
        for subModel in self.submodels:
            subModel.updateLocalCellState(self)
        
    def calcMass(self):
        for comp in self.compartments:
            if comp.id == 'c':
                iCellComp = comp.index
    
        mass = 0.
        for species in self.species:
            if species.molecularWeight is not None:
                mass += self.speciesCounts[species.index, iCellComp] * species.molecularWeight
        mass /= N_AVOGADRO
        
        self.mass = mass
        self.dryWeight = self.fractionDryWeight * mass
    
    def calcVolume(self):
        self.volume = self.mass / self.density
        
    def setComponentIndices(self):
        for index, obj in enumerate(self.submodels):
            obj.index = index
        for index, obj in enumerate(self.compartments):
            obj.index = index
        for index, obj in enumerate(self.species):
            obj.index = index
        for index, obj in enumerate(self.reactions):
            obj.index = index
        for index, obj in enumerate(self.parameters):
            obj.index = index
        for index, obj in enumerate(self.references):
            obj.index = index
        
    #get species counts as dictionary
    def getSpeciesCountsDict(self):           
        speciesCountsDict = {}
        for species in self.species:
            for compartment in self.compartments:
                speciesCountsDict['%s[%s]' % (species.id, compartment.id)] = self.speciesCounts[species.index, compartment.index]                
        return speciesCountsDict
    
    #set species counts for dictionary
    def setSpeciesCountsDict(self, speciesCountsDict):
        for species in self.species:
            for compartment in self.compartments:
                self.speciesCounts[species.index, compartment.index] = speciesCountsDict['%s[%s]' % (species.id, compartment.id)]
                
    #get species concentrations
    def getSpeciesConcentrations(self):
        return self.speciesCounts / self.getSpeciesVolumes() / N_AVOGADRO
        
    #get species concentrations
    def getSpeciesConcentrationsDict(self):
        concs = self.getSpeciesConcentrations()
        speciesConcsDict = {}
        for species in self.species:
            for compartment in self.compartments:
                speciesConcsDict['%s[%s]' % (species.id, compartment.id)] = concs[species.index, compartment.index]                
        return speciesConcsDict
    
    #get container volumes for each species
    def getSpeciesVolumes(self):
        cellComp = self.getComponentById('c', self.compartments)
        extracellularComp = self.getComponentById('e', self.compartments)
        
        volumes = np.zeros((len(self.species), len(self.compartments)))
        volumes[:, cellComp.index] = self.volume
        volumes[:, extracellularComp.index] = self.extracellularVolume
        return volumes
        
    #get species counts as dictionary
    def getSpeciesVolumesDict(self):  
        volumes = self.getSpeciesVolumes()         
        volumesDict = {}
        for species in self.species:
            for compartment in self.compartments:
                volumesDict['%s[%s]' % (species.id, compartment.id)] = volumes[species.index, compartment.index]                
        return volumesDict
    
    #get total RNA number        
    def getTotalRnaCount(self):
        cellComp = self.getComponentById('c', self.compartments)
        tot = 0
        for species in self.species:
            if species.type == 'RNA':
               tot += self.speciesCounts[species.index, cellComp.index]
        return tot
         
    #get total protein copy number
    def getTotalProteinCount(self):
        cellComp = self.getComponentById('c', self.compartments)
        tot = 0
        for species in self.species:
            if species.type == 'Protein':
               tot += self.speciesCounts[species.index, cellComp.index]
        return tot
    
    def getComponentById(self, id, components = None):
        if not components:
            components = chain(self.submodels, self.compartments, self.species, self.reactions, self.parameters, self.references)
            
        for component in components:
            if component.id == id:
                return component
               
#Represents a submodel
class Submodel:
    index = None
    id = ''
    name = ''
    algorithm = ''
    
    reactions = []
    species = []
    parameters = []
    
    speciesCounts = np.zeros(0)
    volume = np.zeros(0)
    extracellularVolume = np.zeros(0)
    
    def __init__(self, id = '', name = '', reactions = [], species = []):
        self.id = id
        self.name = name
        self.reactions = reactions
        self.species = species
        
    def setupSimulation(self):
        #initialize species counts dictionary
        self.speciesCounts = {}
        for species in self.species:
            self.speciesCounts[species.id] = 0
        
    #sets local species counts from global species counts
    def updateLocalCellState(self, model):
        for species in self.species:
            self.speciesCounts[species.id] = model.speciesCounts[species.species.index, species.compartment.index]
        self.volume = model.volume
        self.extracellularVolume = model.extracellularVolume
    
    #sets global species counts from local species counts 
    def updateGlobalCellState(self, model):
        for species in self.species:
            model.speciesCounts[species.species.index, species.compartment.index] = self.speciesCounts[species.id]
            
    #get species concentrations
    def getSpeciesConcentrations(self):
        volumes = self.getSpeciesVolumes()
        concs = {}
        for species in self.species:
            concs[species.id] = self.speciesCounts[species.id] / volumes[species.id] / N_AVOGADRO
        return concs
        
    #get container volumes for each species
    def getSpeciesVolumes(self):
        volumes = {}
        for species in self.species:
            if species.compartment.id == 'c':
                volumes[species.id] = self.volume
            else:
                volumes[species.id] = self.extracellularVolume
        return volumes
        
    #calculate reaction rates
    @staticmethod
    def calcReactionRates(reactions, speciesConcentrations):
        rates = np.full(len(reactions), np.nan)
        for iRxn, rxn in enumerate(reactions):          
            if rxn.rateLaw:
                rates[iRxn] = eval(rxn.rateLaw.transcoded, {}, {'speciesConcentrations': speciesConcentrations, 'Vmax': rxn.vmax, 'Km': rxn.km})
        return rates
               
    #update species counts based on a reaction
    @staticmethod
    def executeReaction(speciesCounts, reaction):
        for part in reaction.participants:
            speciesCounts[part.id] += part.coefficient
        return speciesCounts
    
    def getComponentById(self, id, components = None):
        if not components:
            components = chain(self.species, self.reactions, self.parameters)
        
        for component in components:
            if component.id == id:
                return component
        
#Represents an FBA submodel
class FbaSubmodel(Submodel):
    metabolismProductionReaction = None 
    exchangedSpecies = None
    
    cobraModel = None
    thermodynamicBounds = None
    exchangeRateBounds = None
    
    defaultFbaBound = 1e15
    
    dryWeight = np.nan
    reactionFluxes = np.zeros(0)
    growth = np.nan
    
    def __init__(self, *args, **kwargs):        
        Submodel.__init__(self, *args, **kwargs)
        self.algorithm = 'FBA'
        
    def setupSimulation(self):
        '''setup reaction participant, enzyme counts matrices'''
        Submodel.setupSimulation(self)        
                
        '''Setup FBA'''
        cobraModel = CobraModel(self.id)
        self.cobraModel = cobraModel
            
        #setup metabolites
        cbMets = []
        for species in self.species:
            cbMets.append(CobraMetabolite(id = species.id, name = species.name))
        cobraModel.add_metabolites(cbMets)
        
        #setup reactions
        for rxn in self.reactions:            
            cbRxn = CobraReaction(
                id = rxn.id,
                name = rxn.name,
                lower_bound = -self.defaultFbaBound if rxn.reversible else 0,
                upper_bound =  self.defaultFbaBound,
                objective_coefficient = 1 if rxn.id == 'MetabolismProduction' else 0,
                )
            cobraModel.add_reaction(cbRxn)

            cbMets = {}
            for part in rxn.participants:
                cbMets[part.id] = part.coefficient
            cbRxn.add_metabolites(cbMets)            
        
        #add external exchange reactions
        self.exchangedSpecies = []
        for species in self.species:
            if species.compartment.id == 'e':                
                cbRxn = CobraReaction(
                    id = '%sEx' % species.species.id,
                    name = '%s exchange' % species.species.name,
                    lower_bound = -self.defaultFbaBound,
                    upper_bound =  self.defaultFbaBound,
                    objective_coefficient = 0,
                    )
                cobraModel.add_reaction(cbRxn)
                cbRxn.add_metabolites({species.id: 1})
                
                self.exchangedSpecies.append(ExchangedSpecies(id = species.id, reactionIndex = cobraModel.reactions.index(cbRxn)))
        
        #add biomass exchange reaction
        cbRxn = CobraReaction(
            id = 'BiomassEx',
            name = 'Biomass exchange',
            lower_bound = 0,
            upper_bound = self.defaultFbaBound,
            objective_coefficient = 0,
            )
        cobraModel.add_reaction(cbRxn)
        cbRxn.add_metabolites({'Biomass[c]': -1})
        
        '''Bounds'''
        #thermodynamic       
        arrayCobraModel = cobraModel.to_array_based_model()
        self.thermodynamicBounds = {
            'lower': np.array(arrayCobraModel.lower_bounds.tolist()),
            'upper': np.array(arrayCobraModel.upper_bounds.tolist()),
            }
        
        #exchange reactions
        carbonExRate = self.getComponentById('carbonExchangeRate', self.parameters).value
        nonCarbonExRate = self.getComponentById('nonCarbonExchangeRate', self.parameters).value
        self.exchangeRateBounds = {
            'lower': np.full(len(cobraModel.reactions), -np.nan),
            'upper': np.full(len(cobraModel.reactions),  np.nan),
            }
        for exSpecies in self.exchangedSpecies:
            if self.getComponentById(exSpecies.id, self.species).species.containsCarbon():
                self.exchangeRateBounds['lower'][exSpecies.reactionIndex] = -carbonExRate
                self.exchangeRateBounds['upper'][exSpecies.reactionIndex] =  carbonExRate
            else:
                self.exchangeRateBounds['lower'][exSpecies.reactionIndex] = -nonCarbonExRate
                self.exchangeRateBounds['upper'][exSpecies.reactionIndex] =  nonCarbonExRate
            
        '''Setup reactions'''
        self.metabolismProductionReaction = {
            'index': cobraModel.reactions.index(cobraModel.reactions.get_by_id('MetabolismProduction')),
            'reaction': self.getComponentById('MetabolismProduction', self.reactions),
            }
            
    def updateLocalCellState(self, model):
        Submodel.updateLocalCellState(self, model)
        self.dryWeight = model.dryWeight
        
    def updateGlobalCellState(self, model):
        Submodel.updateGlobalCellState(self, model)
        model.growth = self.growth
                        
    def calcReactionFluxes(self, timeStep = 1):
        '''calc and set bounds'''
        bounds = self.calcReactionBounds(timeStep)
        arrCbModel = self.cobraModel.to_array_based_model()
        arrCbModel.lower_bounds = bounds['lower']
        arrCbModel.upper_bounds = bounds['upper']
        
        '''calculate growth rate'''
        self.cobraModel.optimize()
        
        self.reactionFluxes = self.cobraModel.solution.x
        self.growth = self.reactionFluxes[self.metabolismProductionReaction['index']] #fraction cell/s
        
    def updateMetabolites(self, timeStep = 1):
        #biomass production
        for part in self.metabolismProductionReaction['reaction'].participants:
            self.speciesCounts[part.id] -= self.growth * part.coefficient * timeStep
        
        #external nutrients
        for exSpecies in self.exchangedSpecies:
            self.speciesCounts[exSpecies.id] += self.reactionFluxes[exSpecies.reactionIndex] * timeStep
        
    def calcReactionBounds(self,  timeStep = 1):
        #thermodynamics
        lowerBounds = self.thermodynamicBounds['lower'].copy()
        upperBounds = self.thermodynamicBounds['upper'].copy()
        
        #rate laws
        upperBounds[0:len(self.reactions)] = util.nanminimum(
            upperBounds[0:len(self.reactions)], 
            self.calcReactionRates(self.reactions, self.getSpeciesConcentrations()) * self.volume * N_AVOGADRO,
            )
        
        #external nutrients availability
        for exSpecies in self.exchangedSpecies:
            upperBounds[exSpecies.reactionIndex] = max(0, np.minimum(upperBounds[exSpecies.reactionIndex], self.speciesCounts[exSpecies.id]) / timeStep)
        
        #exchange bounds
        lowerBounds = util.nanminimum(lowerBounds, self.dryWeight / 3600 * N_AVOGADRO * 1e-3 * self.exchangeRateBounds['lower'])
        upperBounds = util.nanminimum(upperBounds, self.dryWeight / 3600 * N_AVOGADRO * 1e-3 * self.exchangeRateBounds['upper'])
        
        #return
        return {'lower': lowerBounds, 'upper': upperBounds}
        
#Represents an SSA submodel
class SsaSubmodel(Submodel):
    def __init__(self, *args, **kwargs):
        Submodel.__init__(self, *args, **kwargs)
        self.algorithm = 'SSA'
        
    def setupSimulation(self):
        Submodel.setupSimulation(self)
            
    @staticmethod
    def stochasticSimulationAlgorithm(speciesCounts, speciesVolumes, reactions, volume, timeMax):
        if len(reactions) >= 1 and not isinstance(reactions[0], list):
            reactions = [reactions]
            
        nSubmodels = len(reactions)

        time = 0
        while time < timeMax:
            #calculate concentrations
            speciesConcentrations = {}
            for id, cnt in speciesCounts.iteritems():
                speciesConcentrations[id] = speciesCounts[id] / speciesVolumes[id] / N_AVOGADRO
        
            #calculate propensities
            totalPropensities = np.zeros(nSubmodels)
            reactionPropensities = []
            for iSubmodel in range(nSubmodels):
                p = np.maximum(0, Submodel.calcReactionRates(reactions[iSubmodel], speciesConcentrations) * volume * N_AVOGADRO)
                totalPropensities[iSubmodel] = np.sum(p)
                reactionPropensities.append(p)
            
            #Select time to next reaction from exponential distribution
            dt = random.exponential(1/np.sum(totalPropensities))
            if time + dt > timeMax:
                if random.rand() > (timeMax - time) / dt:                
                    break
                else:
                    dt = timeMax - time
            
            #Select next reaction
            iSubmodel = random.choice(nSubmodels, p = totalPropensities / np.sum(totalPropensities))                    
            iRxn = random.choice(len(reactionPropensities[iSubmodel]), p = reactionPropensities[iSubmodel] / totalPropensities[iSubmodel])

            #update time and execute reaction
            time += dt
            speciesCounts = Submodel.executeReaction(speciesCounts, reactions[iSubmodel][iRxn])
                
        return speciesCounts
        
#Represents a compartment
class Compartment:
    index = None
    id = ''
    name = ''
    initialVolume = None
    comments = ''
    
    def __init__(self, id = '', name = '', initialVolume = None, comments = ''):
        self.id = id
        self.name = name
        self.initialVolume = initialVolume
        self.comments = comments
        
#Represents a species
class Species:
    index = None
    id = ''
    name = ''
    structure = ''
    empiricalFormula = ''
    molecularWeight = None
    charge = None
    type = ''
    concentrations = []
    crossRefs = []
    comments = ''
    
    def __init__(self, id = '', name = '', structure = '', empiricalFormula = '', molecularWeight = None, 
        charge = None, type = '', concentrations = [], crossRefs = [], comments = ''):
        
        self.id = id    
        self.name = name
        self.structure = structure
        self.empiricalFormula = empiricalFormula
        self.molecularWeight = molecularWeight
        self.charge = charge
        self.type = type
        self.concentrations = concentrations
        self.crossRefs = crossRefs
        
    def containsCarbon(self):
        if self.empiricalFormula:
            return self.empiricalFormula.upper().find('C') != -1
        return False

#Represents a reaction
class Reaction:
    index = None
    id = ''
    name = ''
    submodel = ''
    reversible = None
    participants = []
    enzyme = ''
    rateLaw = None
    vmax = None
    km = None
    crossRefs = []
    comments = ''

    def __init__(self, id = '', name = '', submodel = '', reversible = None, participants = [], 
        enzyme = '', rateLaw = '', vmax = None, km = None, crossRefs = [], comments = ''):
        
        if vmax:
            vmax = float(vmax)
        if km:
            km = float(km)
        
        self.id = id    
        self.name = name
        self.submodel = submodel
        self.reversible = reversible
        self.participants = participants
        self.enzyme = enzyme
        self.rateLaw = rateLaw
        self.vmax = vmax
        self.km = km
        self.crossRefs = crossRefs
        self.comments = comments
        
    #convert rate law to python        
    def getStoichiometryString(self):
        globalComp = self.participants[0].compartment
        for part in self.participants:
            if part.compartment != globalComp:
                globalComp = None
                break
        
        lhs = []
        rhs = []
        for part in self.participants:
            if part.coefficient < 0:
                partStr = ''
                if part.coefficient != -1:
                    if math.ceil(part.coefficient) == part.coefficient: 
                        partStr += '(%d) ' % -part.coefficient
                    else:
                        partStr += '(%e) ' % -part.coefficient
                partStr += part.species.id
                if globalComp is None:
                    partStr += '[%s]' % part.compartment.id
                lhs.append(partStr)
            else:
                partStr = ''
                if part.coefficient != 1:
                    if math.ceil(part.coefficient) == part.coefficient: 
                        partStr += '(%d) ' % part.coefficient
                    else:
                        partStr += '(%e) ' % part.coefficient
                partStr += part.species.id
                if globalComp is None:
                    partStr += '[%s]' % part.compartment.id
                rhs.append(partStr)
            
        stoichStr = ''
        if globalComp is not None:
            stoichStr += '[%s]: ' % globalComp.id
        stoichStr += '%s %s==> %s' % (' + '.join(lhs), '<' if self.reversible else '', ' + '.join(rhs))
        
        return stoichStr
        
#Represents a model parameter
class Parameter:
    index = None
    id = ''
    name = ''
    submodel = None
    value = None
    units = ''
    comments = ''
    
    def __init__(self, id = '', name = '', submodel = '', value = None, units = '', comments = ''):
        self.id = id
        self.name = name
        self.submodel = submodel
        self.value = value
        self.units = units
        self.comments = comments
            
#Represents a reference
class Reference:
    index = None
    id = ''
    name = ''
    crossRefs = []
    comments = ''
    
    def __init__(self, id = '', name = '', crossRefs = [], comments = ''):
        self.id = id
        self.name = name
        self.crossRefs = crossRefs
        self.comments = comments

#Represents a concentration in a compartment
class Concentration:
    compartment = ''
    value = None
    
    def __init__(self, compartment = '', value = None):
        self.compartment = compartment
        self.value = value
    
#Represents a participant in a submodel
class SpeciesCompartment:
    index = None    
    species = ''
    compartment = ''
    
    id = ''
    name = ''
    
    def __init__(self, index = None, species = '', compartment = ''):
        self.index = index
        self.species = species
        self.compartment = compartment    
        
    def calcIdName(self):
        self.id = '%s[%s]' % (self.species.id, self.compartment.id)
        self.name = '%s (%s)' % (self.species.name, self.compartment.name)
      
#Represents an external 
class ExchangedSpecies:
    id = ''
    reactionIndex = None
    
    def __init__(self, id = '', reactionIndex = None):
        self.id = id
        self.reactionIndex = reactionIndex
    
#Represents a participant in a reaction
class ReactionParticipant:
    species = ''
    compartment = ''
    coefficient = None
    
    id = ''
    name = ''
    
    def __init__(self, species = '', compartment = '', coefficient = None):
        self.species = species
        self.compartment = compartment    
        self.coefficient = coefficient
        
    def calcIdName(self):
        self.id = '%s[%s]' % (self.species.id, self.compartment.id)
        self.name = '%s (%s)' % (self.species.name, self.compartment.name)

#Represents a rate law
class RateLaw:
    native = ''
    transcoded = ''
    
    def __init__(self, native = ''):
        self.native = native or ''
        
    #get modifiers of rate law
    def getModifiers(self, species, compartments):
        modifiers = []        
        for spec in species:
            for comp in compartments:
                id = '%s[%s]' % (spec.id, comp.id)
                if self.native.find(id) != -1:
                    modifiers.append(id)
        return modifiers
        
    #transcoded for python
    def transcode(self, species, compartments):
        self.transcoded = self.native
        
        for spec in species:
            for comp in compartments:
                id = '%s[%s]' % (spec.id, comp.id)
                self.transcoded = self.transcoded.replace(id, "speciesConcentrations['%s']" % id)
        
#Represents a cross reference to an external database
class CrossReference:
    source = ''
    id = ''    

    def __init__(self, source = '', id = ''):
        self.source = source
        self.id = id

#Reads model from Excel file into a Python object
def getModelFromExcel(filename):
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", "Discarded range with reserved name", UserWarning)
        wb = load_workbook(filename = filename)

    #initialize model object
    model = Model()

    '''Read details from Excel'''
    #submodels
    ws = wb['Submodels']
    for iRow in range(2, ws.max_row + 1):
        id = ws.cell(row = iRow, column = 1).value
        name = ws.cell(row = iRow, column = 2).value
        algorithm = ws.cell(row = iRow, column = 3).value
        if algorithm == 'FBA':
            subModel = FbaSubmodel(id = id, name = name)
        elif algorithm == 'SSA':
            subModel = SsaSubmodel(id = id, name = name)
        else:
            raise Exception('Undefined algorithm "%s" for submodel "%s"' % (algorithm, id))
        model.submodels.append(subModel)
            
    #compartments
    ws = wb['Compartments']
    for iRow in range(2, ws.max_row + 1):
        model.compartments.append(Compartment(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            initialVolume = float(ws.cell(row = iRow, column = 3).value),
            comments = ws.cell(row = iRow, column = 4).value,
            ))
    
    #species
    ws = wb['Species']
    for iRow in range(2, ws.max_row + 1):
        mwStr = ws.cell(row = iRow, column = 5).value
        if mwStr:
            mw = float(mwStr)
        else:
            mw = None
            
        chargeStr = ws.cell(row = iRow, column = 6).value
        if chargeStr:
            charge = float(chargeStr)
        else:
            charge = None
    
        model.species.append(Species(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            structure = ws.cell(row = iRow, column = 3).value,
            empiricalFormula = ws.cell(row = iRow, column = 4).value,
            molecularWeight = mw,
            charge = charge,
            type = ws.cell(row = iRow, column = 7).value,
            concentrations = [
                Concentration(compartment = 'c', value = float(ws.cell(row = iRow, column = 8).value or 0)),
                Concentration(compartment = 'e', value = float(ws.cell(row = iRow, column = 9).value or 0)),
                ],
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 10).value, 
                    id = ws.cell(row = iRow, column = 11).value,
                    ),
                ],
            comments = ws.cell(row = iRow, column = 12).value,
            ))
            
    #reactions
    ws = wb['Reactions']
        
    for iRow in range(2, ws.max_row + 1):
        stoichiometry = parseStoichiometry(ws.cell(row = iRow, column = 4).value)
        
        rateLawStr = ws.cell(row = iRow, column = 6).value
        if rateLawStr:
            rateLaw = RateLaw(rateLawStr)
        else:
            rateLaw = None
        
        model.reactions.append(Reaction(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            submodel = ws.cell(row = iRow, column = 3).value,
            reversible = stoichiometry['reversible'],
            participants = stoichiometry['participants'],
            enzyme = ws.cell(row = iRow, column = 5).value,
            rateLaw = rateLaw,
            vmax = ws.cell(row = iRow, column = 7).value,
            km = ws.cell(row = iRow, column = 8).value,
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 9).value, 
                    id = ws.cell(row = iRow, column = 10).value,
                    ),
                ],
            comments = ws.cell(row = iRow, column = 11).value,
            ))
            
    #parameters
    ws = wb['Parameters']
    for iRow in range(2, ws.max_row + 1):
        model.parameters.append(Parameter(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            submodel = ws.cell(row = iRow, column = 3).value,
            value = float(ws.cell(row = iRow, column = 4).value),
            units = ws.cell(row = iRow, column = 5).value,
			comments = ws.cell(row = iRow, column = 6).value,
            ))
            
    #references
    ws = wb['References']
    for iRow in range(2, ws.max_row + 1):
        model.references.append(Reference(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 3).value, 
                    id = ws.cell(row = iRow, column = 4).value,
                    ),
                ],
            comments = ws.cell(row = iRow, column = 5).value,
            ))
            
    '''set component indices'''
    model.setComponentIndices()
            
    '''deserialize references'''
    undefinedComponents = []
    
    #species concentration 
    for species in model.species:
        for conc in species.concentrations:
            id = conc.compartment            
            obj = model.getComponentById(id, model.compartments)
            if id and obj is None:
                undefinedComponents.append(id)
            conc.compartment = obj                
            
    #reaction submodel, participant species, participant compartments, enzymes
    for reaction in model.reactions:
        id = reaction.submodel        
        obj = model.getComponentById(id, model.submodels)
        if id and obj is None:
            undefinedComponents.append(id)
        reaction.submodel = obj
        
        for part in reaction.participants:
            id = part.species            
            obj = model.getComponentById(id, model.species)
            if id and obj is None:               
                undefinedComponents.append(id)
            part.species = obj
                
            id = part.compartment            
            obj = model.getComponentById(id, model.compartments)
            if id and obj is None:
                undefinedComponents.append(id)
            part.compartment = obj
            
            part.calcIdName()
        
        id = reaction.enzyme
        obj = model.getComponentById(id, model.species)
        if id and obj is None:
            undefinedComponents.append(id)
        reaction.enzyme = obj

    #parameter submodels
    for param in model.parameters:
        id = param.submodel
        if id:
            obj = model.getComponentById(id, model.submodels)
            if obj is None:
                undefinedComponents.append(id)
            param.submodel = obj

    if len(undefinedComponents) > 0:
        undefinedComponents = list(set(undefinedComponents))
        undefinedComponents.sort()
        raise Exception('Undefined components:\n- %s' % ('\n- '.join(undefinedComponents)))
        
    ''' Assemble back rerferences'''
    for subModel in model.submodels:
        subModel.reactions = []
        subModel.species = []
        subModel.parameters = []
    for rxn in model.reactions:
        rxn.submodel.reactions.append(rxn)
        for part in rxn.participants:
            rxn.submodel.species.append('%s[%s]' % (part.species.id, part.compartment.id))
        if rxn.enzyme:
            rxn.submodel.species.append('%s[%s]' % (rxn.enzyme.id, 'c'))
        if rxn.rateLaw:
            rxn.submodel.species += rxn.rateLaw.getModifiers(model.species, model.compartments)
    
    for param in model.parameters:
        if param.submodel:
            param.submodel.parameters.append(param)
            
    for subModel in model.submodels:
        speciesStrArr = list(set(subModel.species))
        speciesStrArr.sort()
        subModel.species = []
        for index, speciesStr in enumerate(speciesStrArr):
            speciesId, compId = speciesStr.split('[')
            compId = compId[0:-1]
            speciesComp = SpeciesCompartment(
                index = index,
                species = model.getComponentById(speciesId, model.species),
                compartment = model.getComponentById(compId, model.compartments),
                )
            speciesComp.calcIdName()
            subModel.species.append(speciesComp)
            
    '''Transcode rate laws'''
    for rxn in model.reactions:
        if rxn.rateLaw:
            rxn.rateLaw.transcode(model.species, model.compartments)
        
    '''Prepare submodels for computation'''
    model.setupSimulation()
        
    '''Return'''
    return model
    
#Parse a string representing the stoichiometry of a reaction into a Python object
def parseStoichiometry(rxnStr):
    #Split stoichiometry in to global compartment, left-hand side, right-hand side, reversibility indictor
	rxnMatch = re.match('(?P<compartment>\[([a-z])\]: )?(?P<lhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?) (?P<direction>[<]?)==> (?P<rhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?)', rxnStr, flags=re.I)
	if rxnMatch is None:
		raise Exception('Invalid stoichiometry: %s' % rxnStr)
		
    #Determine reversiblity
	rxnDict = rxnMatch.groupdict()
	reversible = rxnDict['direction'] == '<'
	
    #Determine if global compartment for reaction was specified
	if rxnDict['compartment'] is None:
		globalComp = None
	else:
		globalComp = re.match('\[(?P<compartment>[a-z])\]', rxnDict['compartment'], flags=re.I).groupdict()['compartment']
	
    #initialize array of reaction participants
	participants = []
	
    #Parse left-hand side
	for rxnPartStr in rxnDict['lhs'].split(' + '):
		rxnPartDict = re.match('(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxnPartStr, flags=re.I).groupdict()
				
		species = rxnPartDict['species']
		compartment = rxnPartDict['compartment'] or globalComp
		coefficient = float(rxnPartDict['coefficient'] or 1)
	
		participants.append(ReactionParticipant(
			species = species,
			compartment = compartment,
			coefficient = -coefficient,
			))
			
    #Parse right-hand side
	for rxnPartStr in rxnDict['rhs'].split(' + '):
		rxnPartDict = re.match('(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxnPartStr, flags=re.I).groupdict()
				
		species = rxnPartDict['species']
		compartment = rxnPartDict['compartment'] or globalComp
		coefficient = float(rxnPartDict['coefficient'] or 1)
	
		participants.append(ReactionParticipant(
			species = species,
			compartment = compartment,
			coefficient = coefficient,
			))

	return {
		'reversible': reversible,
		'participants': participants,
		}
